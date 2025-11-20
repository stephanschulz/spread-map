#!/usr/bin/env python3
"""
Web GUI for Universe Map Generator
Allows customization of grid dimensions and exports Excel files
"""

from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import csv

app = Flask(__name__)

def generate_universe_colors(total_universes):
    """Generate distinct colors with 20% transparency for any number of universes."""
    # Extended color palette with more distinct colors
    base_colors = [
        'CCFF6B6B', 'CC4ECDC4', 'CCFFE66D', 'CC95E1D3', 'CCFF8B94', 'CCA8E6CF',
        'CCFFD93D', 'CC6BCF7F', 'CCFFA07A', 'CC87CEEB', 'CCDDA15E', 'CCB4A7D6',
        'CCFFC0CB', 'CC98D8C8', 'CCFFD700', 'CC9370DB', 'CCFFA500', 'CC20B2AA',
        'CCFF69B4', 'CC7FFFD4', 'CCFFDAB9', 'CC8FBC8F', 'CCFFC1CC', 'CCAFEEEE',
        'CCFFB6C1', 'CC00CED1', 'CCFFA07A', 'CC9ACD32', 'CCFF7F50', 'CC48D1CC',
        'CCFFD700', 'CC9932CC', 'CCFF8C00', 'CC00FA9A', 'CCDA70D6', 'CC32CD32',
    ]
    
    colors = {}
    for i in range(total_universes):
        # Cycle through colors if we have more universes than colors
        color_idx = i % len(base_colors)
        colors[i + 1] = base_colors[color_idx]
    
    return colors

def generate_cell_label(row_in_universe, col_in_universe, universe_num):
    """Generate a cell label with line break: N\nU#-B#\n"""
    col_label = f"{col_in_universe + 1}"  # 1-20 across the row
    universe_label = f"U{universe_num}"
    b_num = row_in_universe  # B1, B2, etc. for each row
    b_label = f"B{b_num}"
    return f"{col_label}\n{universe_label}-{b_label}\n"

def generate_map_data(cols_per_universe, rows_per_universe, universes_horizontal=6, universes_vertical=6):
    """Generate the map grid data."""
    total_rows = rows_per_universe * universes_vertical
    total_cols = cols_per_universe * universes_horizontal
    
    grid = []
    
    # Add header row
    header_row = ['']
    for col_idx in range(total_cols):
        header_row.append(f"C{col_idx + 1}")
    grid.append(header_row)
    
    # Generate data rows
    for row_idx in range(total_rows):
        row_data = []
        row_data.append(f"R{row_idx + 1}")
        
        universe_row = row_idx // rows_per_universe
        row_in_universe = (row_idx % rows_per_universe) + 1
        
        for col_idx in range(total_cols):
            universe_col = col_idx // cols_per_universe
            col_in_universe = col_idx % cols_per_universe
            universe_num = (universe_row * universes_horizontal) + universe_col + 1
            cell_label = generate_cell_label(row_in_universe, col_in_universe, universe_num)
            row_data.append(cell_label)
        
        grid.append(row_data)
    
    return grid

def create_excel_file(grid, total_universes):
    """Create Excel file from grid data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Universe Map"
    
    colors = generate_universe_colors(total_universes)
    
    # Styles
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    center_align = Alignment(horizontal='center', vertical='center')
    top_center_align_wrap = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    light_grey_side = Side(style='thin', color='D3D3D3')
    light_grey_border = Border(left=light_grey_side, right=light_grey_side, 
                               top=light_grey_side, bottom=light_grey_side)
    
    # Write data
    for row_idx, row_data in enumerate(grid, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = light_grey_border
            
            if row_idx == 1 or col_idx == 1:
                cell.alignment = center_align
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.alignment = top_center_align_wrap
                cell.font = data_font
                
                if isinstance(cell_value, str) and 'U' in cell_value:
                    try:
                        lines = cell_value.split('\n')
                        if len(lines) >= 2:
                            second_line = lines[1]
                            parts = second_line.split('-')
                            universe_str = parts[0]
                            universe_num = int(universe_str[1:])
                            
                            if universe_num in colors:
                                cell.fill = PatternFill(start_color=colors[universe_num], 
                                                       end_color=colors[universe_num], 
                                                       fill_type='solid')
                    except (IndexError, ValueError):
                        pass
    
    # Set square cell dimensions
    row_height = 50
    col_width = 7.5
    
    for col_idx in range(1, len(grid[0]) + 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = col_width
    
    for row_idx in range(1, len(grid) + 1):
        ws.row_dimensions[row_idx].height = row_height
    
    return wb

@app.route('/')
def index():
    """Main page with form."""
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    """Generate and download the Excel file."""
    try:
        # Get parameters from form
        cols = int(request.form.get('cols', 20))
        rows = int(request.form.get('rows', 13))
        universes_h = int(request.form.get('universes_h', 6))
        universes_v = int(request.form.get('universes_v', 6))
        
        # Validate inputs
        if not (1 <= cols <= 50 and 1 <= rows <= 50):
            return jsonify({'error': 'Rows and columns must be between 1 and 50'}), 400
        if not (1 <= universes_h <= 10 and 1 <= universes_v <= 10):
            return jsonify({'error': 'Universe counts must be between 1 and 10'}), 400
        
        # Generate data
        total_universes = universes_h * universes_v
        grid = generate_map_data(cols, rows, universes_h, universes_v)
        wb = create_excel_file(grid, total_universes)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='universe_map.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/preview', methods=['POST'])
def preview():
    """Get preview info about the generated map."""
    try:
        cols = int(request.form.get('cols', 20))
        rows = int(request.form.get('rows', 13))
        universes_h = int(request.form.get('universes_h', 6))
        universes_v = int(request.form.get('universes_v', 6))
        
        total_cols = cols * universes_h
        total_rows = rows * universes_v
        total_universes = universes_h * universes_v
        
        return jsonify({
            'total_cols': total_cols,
            'total_rows': total_rows,
            'total_universes': total_universes,
            'grid_size': f"{total_rows + 1} rows × {total_cols + 1} columns (with headers)"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = 5001
    print("\n🚀 Universe Map Generator Web App")
    print("=" * 50)
    print(f"Opening at: http://localhost:{port}")
    print("Press Ctrl+C to stop the server")
    print("=" * 50 + "\n")
    app.run(debug=True, port=port, host='127.0.0.1')

