#!/usr/bin/env python3
"""
Generate a CSV map with multiple universes arranged in a grid.
Each universe is 15 columns x 20 rows.
Also generates an Excel file with colored universes for Google Sheets.
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import os

# Configuration
COLS_PER_UNIVERSE = 20
ROWS_PER_UNIVERSE = 13
UNIVERSES_HORIZONTAL = 6  # U1 through U6 (original + 5 to the right)
UNIVERSES_VERTICAL = 6    # 6 rows of universes (original + 5 downward)

def generate_cell_label(row_in_universe, col_in_universe, universe_num):
    """
    Generate a cell label with line break: N\nU#-B#
    row_in_universe: 1-13
    col_in_universe: 0-19 (maps to B1-B20)
    universe_num: 1-36
    """
    # Row number (1-20 for naming, even though we only have 13 rows)
    # We'll use 1-13 as the actual row numbers
    row_label = f"{row_in_universe}"
    
    # Universe label (U1, U2, etc.)
    universe_label = f"U{universe_num}"
    
    # B label goes from B1 to B20 (one per column in the universe)
    b_num = col_in_universe + 1
    b_label = f"B{b_num}"
    
    # Format with line breaks: row number on first line, U#-B# on second line, empty line at bottom
    return f"{row_label}\n{universe_label}-{b_label}\n"

def generate_map():
    """Generate the complete map with all universes."""
    total_rows = ROWS_PER_UNIVERSE * UNIVERSES_VERTICAL
    total_cols = COLS_PER_UNIVERSE * UNIVERSES_HORIZONTAL
    
    # Create the grid
    grid = []
    
    # Add header row with column labels (C1, C2, C3, ...)
    header_row = ['']  # Empty cell at top-left corner
    for col_idx in range(total_cols):
        header_row.append(f"C{col_idx + 1}")
    grid.append(header_row)
    
    for row_idx in range(total_rows):
        row_data = []
        
        # Add row label at the beginning (R1, R2, R3, ...)
        row_data.append(f"R{row_idx + 1}")
        
        # Determine which universe row we're in (0-5)
        universe_row = row_idx // ROWS_PER_UNIVERSE
        # Row within the current universe (1-20)
        row_in_universe = (row_idx % ROWS_PER_UNIVERSE) + 1
        
        for col_idx in range(total_cols):
            # Determine which universe column we're in (0-5)
            universe_col = col_idx // COLS_PER_UNIVERSE
            # Column within the current universe (0-14)
            col_in_universe = col_idx % COLS_PER_UNIVERSE
            
            # Calculate universe number (1-36)
            # Universe numbering goes left to right, then top to bottom
            universe_num = (universe_row * UNIVERSES_HORIZONTAL) + universe_col + 1
            
            # Generate the cell label
            cell_label = generate_cell_label(row_in_universe, col_in_universe, universe_num)
            row_data.append(cell_label)
        
        grid.append(row_data)
    
    return grid

def generate_universe_colors():
    """
    Generate distinct colors arranged so neighbors are maximally different.
    Uses a strategic color pattern for the 6x6 grid.
    Colors include alpha channel for 20% transparency (80% opacity = CC).
    """
    # 12 highly distinct base colors spread across the color spectrum
    # Format: AARRGGBB where AA=CC (80% opacity for 20% transparency)
    base_colors = [
        'CCFF6B6B',  # Bright Red
        'CC4ECDC4',  # Turquoise
        'CCFFE66D',  # Bright Yellow
        'CC95E1D3',  # Mint
        'CCFF8B94',  # Pink
        'CCA8E6CF',  # Light Green
        'CCFFD93D',  # Golden Yellow
        'CC6BCF7F',  # Green
        'CCFFA07A',  # Light Salmon
        'CC87CEEB',  # Sky Blue
        'CCDDA15E',  # Tan/Brown
        'CCB4A7D6',  # Lavender
    ]
    
    # Pattern for 6x6 grid - each number represents which base color to use
    # This pattern ensures no adjacent cells have the same color
    # Pattern uses modulo-style distribution for maximum contrast
    color_pattern = [
        [0, 3, 1, 4, 2, 5],
        [6, 9, 7, 10, 8, 11],
        [1, 4, 2, 5, 0, 3],
        [7, 10, 8, 11, 6, 9],
        [2, 5, 0, 3, 1, 4],
        [8, 11, 6, 9, 7, 10],
    ]
    
    # Build the color map for all 36 universes
    colors = {}
    for row in range(6):
        for col in range(6):
            universe_num = (row * 6) + col + 1
            color_idx = color_pattern[row][col]
            colors[universe_num] = base_colors[color_idx]
    
    return colors

def save_to_csv(grid, filename='map.csv'):
    """Save the grid to a CSV file."""
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(grid)
    print(f"Generated {filename}")
    print(f"  Total size: {len(grid)} rows x {len(grid[0])} columns (including headers)")

def save_to_excel(grid, filename='map.xlsx', background_image=None, transparent_cells=False):
    """Save the grid to an Excel file with square cells and optional background image."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Universe Map"
    
    # Get universe colors
    colors = generate_universe_colors()
    
    # Header style (bold, gray background)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    header_font = Font(bold=True, size=9)
    data_font = Font(size=9)  # Smaller font for data to fit better in square cells
    center_align = Alignment(horizontal='center', vertical='center')
    top_center_align_wrap = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # Light grey border for all cells
    light_grey_side = Side(style='thin', color='D3D3D3')
    light_grey_border = Border(left=light_grey_side, right=light_grey_side, 
                               top=light_grey_side, bottom=light_grey_side)
    
    # Write data and apply formatting
    for row_idx, row_data in enumerate(grid, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply light grey border to all cells
            cell.border = light_grey_border
            
            # Style header row and column
            if row_idx == 1 or col_idx == 1:
                cell.alignment = center_align
                cell.fill = header_fill
                cell.font = header_font
            else:
                # Apply top-center alignment and font to data cells
                cell.alignment = top_center_align_wrap
                cell.font = data_font
                
                # Only apply colors if not using transparent cells
                if not transparent_cells:
                    # Determine universe number from cell content
                    # Cell format: NN\nU#-B#\n (e.g., 01\nU1-B1\n)
                    if isinstance(cell_value, str) and 'U' in cell_value:
                        try:
                            # Split by newline and get second part
                            lines = cell_value.split('\n')
                            if len(lines) >= 2:
                                # Second line contains U#-B#
                                second_line = lines[1]
                                # Extract universe number (format: U#-B#)
                                parts = second_line.split('-')
                                universe_str = parts[0]  # U1, U2, etc.
                                universe_num = int(universe_str[1:])  # Extract number
                                
                                # Apply color for this universe
                                if universe_num in colors:
                                    cell.fill = PatternFill(start_color=colors[universe_num], 
                                                           end_color=colors[universe_num], 
                                                           fill_type='solid')
                        except (IndexError, ValueError):
                            pass
    
    # Set fixed square cell dimensions
    # Excel uses different units: row height in points, column width in character units
    # To create square cells, we need to match them visually
    row_height = 50  # Height in points
    col_width = 7.5  # Width in character units (approximately matches 50pt height for square cells)
    
    # Set all column widths to create square cells
    for col_idx in range(1, len(grid[0]) + 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = col_width
    
    # Set all row heights to create square cells
    for row_idx in range(1, len(grid) + 1):
        ws.row_dimensions[row_idx].height = row_height
    
    # Add background image AFTER all cells are written so it appears below
    if background_image and os.path.exists(background_image):
        img = Image(background_image)
        # Position the image at cell B2 (start of data area, after headers)
        img.anchor = 'B2'
        ws.add_image(img)
    
    wb.save(filename)
    print(f"Generated {filename}")
    print(f"  Total size: {len(grid)} rows x {len(grid[0])} columns (including headers)")
    if transparent_cells:
        print(f"  Cells are transparent to show background image")
    else:
        print(f"  Each universe has its own color")
    if background_image and os.path.exists(background_image):
        print(f"  Background image added: {background_image}")

if __name__ == '__main__':
    grid = generate_map()
    save_to_csv(grid)
    
    # Check for overlay image (PNG with transparency)
    overlay_image = 'floorplan.png'
    has_overlay = os.path.exists(overlay_image)
    
    if has_overlay:
        # Generate colored version with transparent PNG overlay
        save_to_excel(grid, filename='map.xlsx', 
                     background_image=overlay_image, transparent_cells=False)
        print("\n✨ Transparent PNG overlay added to colored universe grid")
        print("   The colored cells will show through the transparent areas of the PNG")
    else:
        # No overlay image, just generate normal version
        save_to_excel(grid, filename='map.xlsx', transparent_cells=False)
    
    print(f"\nUniverse layout:")
    print(f"  Universes: {UNIVERSES_HORIZONTAL} x {UNIVERSES_VERTICAL} = {UNIVERSES_HORIZONTAL * UNIVERSES_VERTICAL} total")

